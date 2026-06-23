import csv
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from django.utils import timezone


def export_csv(headers, rows):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def export_excel(headers, rows, sheet_name="Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    ws.append(headers)

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    alignment = Alignment(horizontal="left", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Write data rows
    for row in rows:
        formatted_row = []
        for val in row:
            if val is None:
                formatted_row.append("")
            elif isinstance(val, (int, float)):
                formatted_row.append(val)
            else:
                formatted_row.append(str(val))
        ws.append(formatted_row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_pdf(headers, rows, title="Sahāy Admin Report"):
    buffer = BytesIO()

    # Use landscape letter to fit tabular data perfectly
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=12,
    )

    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor("#475569"),
        spaceAfter=20,
    )

    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#1E293B"),
        leading=10,
    )

    header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.white,
        leading=10,
    )

    elements = []

    # Title & Metadata block
    elements.append(Paragraph(title, title_style))
    elements.append(
        Paragraph(
            f"Generated at: {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC')} | Total Records: {len(rows)}",
            meta_style,
        )
    )

    # Format table data with Paragraphs to support text wrapping
    table_data = []
    table_data.append([Paragraph(h, header_style) for h in headers])

    for row in rows:
        table_data.append([Paragraph(str(val) if val is not None else "", cell_style) for val in row])

    # Table styling for clean design
    t = Table(table_data, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),  # Slate-800 header background
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),  # Light gray gridlines
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),  # Zebra striping
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
            ]
        )
    )

    elements.append(t)
    doc.build(elements)
    return buffer.getvalue()
